"""Extract (column code, label) per DORA template from the EBA Annotated Table Layout.

Reads spec/dpm/*Annotated*DORA*.xlsx (gitignored, local). Writes a field map used to
generate the Pydantic domain models. Nothing fabricated — every field traces to the layout.

Run: uv run python scripts/extract_layout.py
"""
from __future__ import annotations
import glob
import json
import pathlib
import re
import openpyxl

CODE = re.compile(r"^\d{4}$")


def extract() -> dict:
    f = glob.glob("spec/dpm/*Annotated*DORA*.xlsx")[0]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    out: dict = {}
    for sn in wb.sheetnames:
        if sn == "TOC":
            continue
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(list(row))
            if i > 20:
                break
        title = next((str(c) for c in rows[0] if c not in (None, "")), sn) if rows else sn
        codes_idx = None
        for i, r in enumerate(rows):
            vals = [str(c).strip() for c in r if c not in (None, "")]
            if vals and sum(1 for v in vals if CODE.match(v)) >= max(1, len(vals) // 2):
                codes_idx = i
                break
        fields = []
        if codes_idx is not None:
            codes = [str(c).strip() for c in rows[codes_idx] if c not in (None, "")]
            labels = [str(c).strip() for c in rows[codes_idx - 1] if c not in (None, "")] if codes_idx else []
            for j, code in enumerate(codes):
                if CODE.match(code):
                    out_label = labels[j] if j < len(labels) else ""
                    fields.append({"code": "c" + code, "label": out_label})
        out[sn] = {"title": title, "fields": fields}
    return out


if __name__ == "__main__":
    data = extract()
    for t, d in data.items():
        print(f"\n## {t}  —  {d['title']}")
        for fld in d["fields"]:
            print(f"  {fld['code']}  {fld['label']}")
    pathlib.Path("spec/_extracted").mkdir(exist_ok=True)
    pathlib.Path("spec/_extracted/template_fields.json").write_text(
        json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"\n[written] spec/_extracted/template_fields.json ({len(data)} templates)")
