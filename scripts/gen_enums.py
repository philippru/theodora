"""Generate src/theodora/domain/enums.py — allowed xBRL dimension members.

Source: spec/dpm/List_of_possible_values*.xlsx (gitignored, local). Each sheet (one per
template) has LIST<dim>/DESC<dim> column pairs holding coded values like 'eba_CU:USD'.
We union them by dimension prefix (eba_CU, eba_GA, eba_CT, ...).

Do NOT hand-edit the output. Regenerate: uv run python scripts/gen_enums.py
"""
from __future__ import annotations

import glob
import pathlib
import re

import openpyxl

OUT = pathlib.Path("src/theodora/domain/enums.py")
VAL = re.compile(r"^(eba_[A-Za-z0-9]+):\S+$")


def main() -> None:
    f = glob.glob("spec/dpm/List_of_possible_values*.xlsx")[0]
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    dims: dict[str, set[str]] = {}
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    m = VAL.match(cell.strip())
                    if m:
                        dims.setdefault(m.group(1), set()).add(cell.strip())

    L: list[str] = [
        '"""Allowed xBRL dimension members — GENERATED from the EBA List of possible values.',
        "",
        "Do NOT hand-edit. Regenerate: uv run python scripts/gen_enums.py",
        '"""',
        "from __future__ import annotations",
        "",
        "",
        "DIMENSIONS: dict[str, frozenset[str]] = {",
    ]
    for dim in sorted(dims):
        L.append(f"    {dim!r}: frozenset({{")
        for v in sorted(dims[dim]):
            L.append(f"        {v!r},")
        L.append("    }),")
    L.append("}")
    L += [
        "",
        "",
        "def dimension_of(value: str) -> str | None:",
        '    """Return the eba_XX prefix of a coded value like "eba_CU:USD", else None."""',
        "    if isinstance(value, str) and value.startswith('eba_') and ':' in value:",
        "        return value.split(':', 1)[0]",
        "    return None",
        "",
    ]
    OUT.write_text("\n".join(L), encoding="utf-8")
    print(f"[written] {OUT} — {len(dims)} dimensions, {sum(len(v) for v in dims.values())} members")


if __name__ == "__main__":
    main()
