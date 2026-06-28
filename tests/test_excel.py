"""Excel adapter tests — round-trip Excel -> domain -> xBRL-CSV -> valid, and label-based headers."""
from __future__ import annotations

from pathlib import Path

import openpyxl

from theodora.generate.builder import build_tables, package_name
from theodora.io.excel import read_workbook, write_workbook
from theodora.io.writer import write_package
from theodora.validation.engine import validate_package

_LEI = "DUMMYLEI123456789012"
_DATE = "2024-12-31"


def test_excel_roundtrip_yields_valid_package(tmp_path: Path) -> None:
    # synthetic valid tables -> Excel -> read back -> xBRL-CSV package -> validate clean
    xlsx = tmp_path / "filled.xlsx"
    write_workbook(xlsx, build_tables(_LEI, _DATE))
    tables = read_workbook(xlsx)
    pkg = write_package(tmp_path / "out", package_name(_LEI, _DATE), tables, _LEI, _DATE)
    assert validate_package(pkg) == []


def test_reads_columns_by_human_label(tmp_path: Path) -> None:
    # A user fills the sheet with the human label, not the xBRL code.
    xlsx = tmp_path / "by_label.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("B_02.01")
    ws.append(["Contractual arrangement reference number"])
    ws.append(["CA-1"])
    wb.save(xlsx)
    tables = read_workbook(xlsx)
    assert tables["B_02.01"] == [{"c0010": "CA-1"}]


def test_reads_columns_by_code(tmp_path: Path) -> None:
    xlsx = tmp_path / "by_code.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("B0201")  # also exercises flexible sheet-name matching
    ws.append(["c0010"])
    ws.append(["CA-7"])
    wb.save(xlsx)
    tables = read_workbook(xlsx)
    assert tables["B_02.01"] == [{"c0010": "CA-7"}]
