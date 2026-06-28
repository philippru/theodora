"""Excel adapter — read a filled RoI workbook into template rows, and write a blank template.

Bridges the real-world workflow: institutions hold their register in Excel, but must submit
xBRL-CSV. `read_workbook` accepts columns identified by EITHER the xBRL code (c0010) OR the
human field label (in any reasonable casing). Sheet names are matched flexibly (B_01.01 /
B0101 / b_01.01). Output is the same `tables` structure the writer consumes.
"""
from __future__ import annotations

import datetime
from pathlib import Path

import openpyxl

from theodora.domain.templates import TEMPLATES
from theodora.validation.engine import _alias_labels, _norm

_SHEET_TO_TID = {tid.upper().replace("_", "").replace(".", ""): tid for tid in TEMPLATES}


def _sheet_tid(name: str) -> str | None:
    return _SHEET_TO_TID.get(name.upper().replace("_", "").replace(".", "").replace(" ", ""))


def _header_map(tid: str) -> dict[str, str]:
    """Accept a column header by xBRL code OR (normalised) label -> alias."""
    m: dict[str, str] = {}
    for alias, label in _alias_labels(tid).items():
        m[alias.lower()] = alias
        if label:
            m[label] = alias
    return m


def _cell(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_workbook(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    tables: dict[str, list[dict]] = {tid: [] for tid in TEMPLATES}
    for sn in wb.sheetnames:
        tid = _sheet_tid(sn)
        if not tid:
            continue
        hmap = _header_map(tid)
        rows = list(wb[sn].iter_rows(values_only=True))
        if not rows:
            continue
        col_alias: dict[int, str] = {}
        for i, h in enumerate(rows[0]):
            if h is None:
                continue
            ht = str(h).strip()
            alias = hmap.get(ht.lower()) or hmap.get(_norm(ht))
            if alias:
                col_alias[i] = alias
        for r in rows[1:]:
            row = {col_alias[i]: _cell(r[i]) for i in col_alias if i < len(r) and _cell(r[i]) != ""}
            if row:
                tables[tid].append(row)
    return tables


def write_workbook(path: Path, tables: dict[str, list[dict]]) -> Path:
    """Write a workbook with one sheet per template (header = xBRL codes). Empty tables -> blank template."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for tid, model in TEMPLATES.items():
        ws = wb.create_sheet(tid)
        headers = [fi.alias for fi in model.model_fields.values() if fi.alias]
        ws.append(headers)
        for row in tables.get(tid, []):
            ws.append([row.get(h, "") for h in headers])
    wb.save(Path(path))
    return Path(path)
